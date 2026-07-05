from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

from .report import write_excel
from .storage import load_state


def _load_any(path: Path) -> dict[str, Any]:
    state = load_state(path)
    if state is None:
        raise FileNotFoundError(path)
    return state


def merge_states(input_paths: list[Path]) -> list[dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = defaultdict(dict)
    for path in input_paths:
        state = _load_any(path)
        env = state.get("environment") or {}
        country = str(env.get("country") or env.get("public_ip") or path.stem or "unknown").replace(" ", "_").lower()
        for result in state.get("results") or []:
            if not isinstance(result, dict):
                continue
            model_id = str(result.get("model_id") or "")
            if not model_id:
                continue
            row = rows[model_id]
            row["model_id"] = model_id
            row[f"{country}_status"] = result.get("test_status", "")
            row[f"{country}_latency_ms"] = result.get("latency_total_ms", "")
            row[f"{country}_http_status"] = result.get("http_status", "")
            row[f"{country}_error_type"] = result.get("error_type", "")
    merged = list(rows.values())
    for row in merged:
        available_regions = [key.removesuffix("_status") for key, value in row.items() if key.endswith("_status") and value == "available"]
        row["available_region_count"] = len(available_regions)
        row["available_regions"] = ",".join(available_regions)
    merged.sort(key=lambda item: (-int(item.get("available_region_count", 0)), str(item.get("model_id", ""))))
    return merged


def write_merge_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    columns: list[str] = ["model_id", "available_region_count", "available_regions"]
    dynamic = sorted({key for row in rows for key in row.keys() if key not in columns})
    columns.extend(dynamic)
    import csv

    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_merge_excel(path: Path, rows: list[dict[str, Any]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    columns: list[str] = ["model_id", "available_region_count", "available_regions"]
    dynamic = sorted({key for row in rows for key in row.keys() if key not in columns})
    columns.extend(dynamic)
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Merge"
    ws.append(columns)
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    for row in rows:
        ws.append([row.get(column, "") for column in columns])
        for cell in ws[ws.max_row]:
            if str(cell.value) == "available":
                cell.fill = green
            elif str(cell.value) == "skipped":
                cell.fill = yellow
            elif cell.column > 3 and str(cell.value):
                cell.fill = red
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:200])
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 10), 60)
    ws.freeze_panes = "A2"
    workbook.save(path)
    return True
