from __future__ import annotations

import csv
import statistics
from pathlib import Path
from typing import Any

RESULT_COLUMNS = [
    "model_id",
    "display_name",
    "test_status",
    "latency_total_ms",
    "http_status",
    "model_type",
    "context_length",
    "max_output_tokens",
    "supports_image_input",
    "supports_coding",
    "supports_reasoning",
    "supports_function_calling",
    "supports_tools",
    "supports_json_mode",
    "supports_streaming",
    "supports_embedding",
    "vector_dimension",
    "provider",
    "endpoint_type",
    "capability_tags",
    "usecase_tags",
    "error_type",
    "error_message",
    "skip_reason",
    "tested_at_utc",
]

ENV_COLUMNS = [
    "tested_at_utc",
    "hostname",
    "fqdn",
    "public_ip",
    "country",
    "region",
    "city",
    "isp",
    "timezone",
    "os",
    "system",
    "release",
    "machine",
    "python_version",
    "proxy_enabled",
    "https_proxy_set",
    "http_proxy_set",
    "no_proxy_set",
    "ip_lookup_error",
    "ip_lookup_error_message",
]


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple, set)):
        return str(value)
    return str(value)


def _safe_number(value: Any, default: float = 999999999.0) -> float:
    if value in (None, "", "unknown", "None"):
        return default
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return default


def _available_sort_key(item: dict[str, Any]) -> tuple[float, float, str]:
    return (
        _safe_number(item.get("latency_total_ms")),
        -_safe_number(item.get("context_length"), default=0.0),
        str(item.get("model_id", "")),
    )


def _all_results_sort_key(item: dict[str, Any]) -> tuple[int, float, str]:
    return (
        0 if item.get("test_status") == "available" else 1,
        _safe_number(item.get("latency_total_ms")),
        str(item.get("model_id", "")),
    )


def _sorted_results(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(results, key=_all_results_sort_key)


def _available_results(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted([item for item in results if item.get("test_status") == "available"], key=_available_sort_key)


def calculate_summary(results: list[dict[str, Any]]) -> dict[str, Any]:
    total = len(results)
    available = sum(1 for item in results if item.get("test_status") == "available")
    skipped = sum(1 for item in results if item.get("test_status") == "skipped")
    failed = total - available - skipped
    latencies = [int(item.get("latency_total_ms") or 0) for item in results if item.get("test_status") == "available" and item.get("latency_total_ms")]
    p50 = int(statistics.median(latencies)) if latencies else ""
    p90 = ""
    if len(latencies) >= 2:
        p90 = int(statistics.quantiles(sorted(latencies), n=10)[8])
    elif len(latencies) == 1:
        p90 = latencies[0]
    return {
        "total_rows": total,
        "available": available,
        "failed": failed,
        "skipped": skipped,
        "available_with_image_input": sum(1 for item in results if item.get("test_status") == "available" and str(item.get("supports_image_input", "")).lower() == "true"),
        "available_with_coding": sum(1 for item in results if item.get("test_status") == "available" and str(item.get("supports_coding", "")).lower() == "true"),
        "available_with_reasoning": sum(1 for item in results if item.get("test_status") == "available" and str(item.get("supports_reasoning", "")).lower() == "true"),
        "available_with_tool_or_function_calling": sum(1 for item in results if item.get("test_status") == "available" and str(item.get("supports_function_calling", "")).lower() == "true"),
        "success_rate": round(available / total, 4) if total else 0,
        "avg_latency_ms": int(sum(latencies) / len(latencies)) if latencies else "",
        "p50_latency_ms": p50,
        "p90_latency_ms": p90,
        "http_429_count": sum(1 for item in results if item.get("http_status") == 429),
        "http_403_count": sum(1 for item in results if item.get("http_status") == 403),
        "http_5xx_count": sum(1 for item in results if isinstance(item.get("http_status"), int) and 500 <= item["http_status"] <= 599),
    }


def write_csv(path: Path, results: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=RESULT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for item in _sorted_results(results):
            writer.writerow({column: _stringify(item.get(column, "")) for column in RESULT_COLUMNS})


def write_excel(path: Path, results: list[dict[str, Any]], summary: dict[str, Any], environment: dict[str, Any]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    summary_ws.append(["metric", "value"])
    for key, value in summary.items():
        summary_ws.append([key, value])

    available_ws = workbook.create_sheet("Available")
    available_ws.append(RESULT_COLUMNS)
    for item in _available_results(results):
        available_ws.append([_stringify(item.get(column, "")) for column in RESULT_COLUMNS])

    model_ws = workbook.create_sheet("All Models")
    model_ws.append(RESULT_COLUMNS)
    fills = {
        "available": PatternFill("solid", fgColor="C6EFCE"),
        "skipped": PatternFill("solid", fgColor="FFEB9C"),
        "timeout": PatternFill("solid", fgColor="FCE4D6"),
        "server_error": PatternFill("solid", fgColor="FCE4D6"),
        "rate_limited": PatternFill("solid", fgColor="FFC7CE"),
        "unauthorized": PatternFill("solid", fgColor="FFC7CE"),
        "forbidden_or_region_block": PatternFill("solid", fgColor="FFC7CE"),
    }
    for item in _sorted_results(results):
        row = [_stringify(item.get(column, "")) for column in RESULT_COLUMNS]
        model_ws.append(row)
        fill = fills.get(str(item.get("test_status", "")))
        if fill:
            for cell in model_ws[model_ws.max_row]:
                cell.fill = fill

    errors_ws = workbook.create_sheet("Errors")
    errors_ws.append(RESULT_COLUMNS)
    for item in results:
        if item.get("test_status") not in {"available", "skipped"}:
            errors_ws.append([_stringify(item.get(column, "")) for column in RESULT_COLUMNS])

    env_ws = workbook.create_sheet("Environment")
    env_ws.append(["key", "value"])
    for key in ENV_COLUMNS:
        env_ws.append([key, _stringify(environment.get(key, ""))])
    for key, value in environment.items():
        if key not in ENV_COLUMNS:
            env_ws.append([key, _stringify(value)])

    for worksheet in workbook.worksheets:
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells[:200])
            width = min(max(max_length + 2, 10), 60)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        worksheet.freeze_panes = "A2"

    workbook.save(path)
    return True


def print_table(results: list[dict[str, Any]], limit: int = 30) -> None:
    try:
        from rich.console import Console
        from rich.table import Table
    except ImportError:
        print("model_id,status,latency_ms,context,max_output,vision,coding,reasoning,tools,error")
        for item in _sorted_results(results)[:limit]:
            print(
                f"{item.get('model_id','')},{item.get('test_status','')},{item.get('latency_total_ms','')},"
                f"{item.get('context_length','')},{item.get('max_output_tokens','')},"
                f"{item.get('supports_image_input','')},{item.get('supports_coding','')},"
                f"{item.get('supports_reasoning','')},{item.get('supports_function_calling','')},"
                f"{str(item.get('error_type',''))[:80]}"
            )
        return

    table = Table(title="NVIDIA Model Probe Results")
    table.add_column("model_id", overflow="fold")
    table.add_column("status")
    table.add_column("latency_ms", justify="right")
    table.add_column("ctx", justify="right")
    table.add_column("max_out", justify="right")
    table.add_column("vision")
    table.add_column("coding")
    table.add_column("reasoning")
    table.add_column("tools")
    table.add_column("error", overflow="fold")
    for item in _sorted_results(results)[:limit]:
        table.add_row(
            str(item.get("model_id", "")),
            str(item.get("test_status", "")),
            str(item.get("latency_total_ms", "")),
            str(item.get("context_length", "")),
            str(item.get("max_output_tokens", "")),
            str(item.get("supports_image_input", "")),
            str(item.get("supports_coding", "")),
            str(item.get("supports_reasoning", "")),
            str(item.get("supports_function_calling", "")),
            str(item.get("error_type", ""))[:80],
        )
    Console().print(table)
